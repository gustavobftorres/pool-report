"""
Data export service for pool metrics.

Exports comprehensive pool analysis data to Excel and CSV formats
for external analysis and RAG model consumption.

Data Structure:
--------------
The exporter creates a unified flat structure with these columns:
- pool_name: Pool name (e.g., "50WETH-50USDC")
- pool_address: Pool address (e.g., "0x3de27...")
- pool_type: "Input Pool" or "Competitor"
- dex: DEX name (e.g., "Balancer", "UniSwap")
- blockchain: Chain name (e.g., "ethereum")
- metric_group: Group name (e.g., "Demand / Usage Metrics")
- metric_name: Metric name (e.g., "Total Traded Volume")
- metric_value: Numeric or string value
- metric_unit: Unit (e.g., "USD", "Count", "%")
- timestamp: Export timestamp

Excel Format:
------------
Sheet 1: "all_metrics" - All pools and all metrics in flat structure
Sheet 2: "summary" - High-level stats per pool
Sheet 3: "anchor_tokens" - Lending market data (if provided)

CSV Format:
----------
Single file with same structure as "all_metrics" sheet
"""
from __future__ import annotations

import pandas as pd
from datetime import datetime, timezone, timedelta
from typing import Any
from pathlib import Path

from services.dune_metrics import METRIC_GROUP_NAMES


# Map parameter change types to display names for Excel export
CHANGE_TYPE_NAMES = {
    "swap_fee": "Swap Fee Change",
    "weights": "Weight Adjustment (LBP)",
    "amp_factor": "Amplification Factor Change",
    "surge_threshold": "Surge Threshold Change",
    "max_surge_fee": "Max Surge Fee Change",
}


class DataExporter:
    """Export pool metrics to spreadsheet formats."""
    
    def __init__(self, export_dir: str = "exports"):
        """
        Initialize exporter.
        
        Args:
            export_dir: Directory to save export files (created if missing)
        """
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(exist_ok=True)
        
    def export_to_excel(
        self,
        results: dict[str, Any],
        anchor_data: dict[str, Any] | None = None,
        anchor_df: Any | None = None,
        filename: str | None = None,
    ) -> str:
        """
        Export metrics pipeline results to Excel file.
        
        Args:
            results: Output from MetricsPipeline.analyze_pool_with_competitors()
            anchor_data: Optional anchor token data from AnchorTokenInfo (summary stats)
            anchor_df: Optional pandas DataFrame with full anchor token data
            filename: Optional custom filename (auto-generated if None)
            
        Returns:
            Path to created Excel file
            
        Example:
            >>> exporter = DataExporter()
            >>> pipeline = MetricsPipeline()
            >>> results = await pipeline.analyze_pool_with_competitors('0x123...')
            >>> excel_path = exporter.export_to_excel(results)
            >>> print(f"Exported to: {excel_path}")
        """
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            pool_addr = results.get("input_pool", {}).get("pool_address", "unknown")[:8]
            
            # Include anchor token in filename if available
            if anchor_data and anchor_data.get("token_symbol"):
                token_symbol = anchor_data["token_symbol"]
                filename = f"PoolReport_{pool_addr}_{token_symbol}_{timestamp}.xlsx"
            else:
                filename = f"PoolReport_{pool_addr}_{timestamp}.xlsx"
        
        filepath = self.export_dir / filename
        
        # Create Excel writer with openpyxl
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Sheet 1: All metrics
            all_metrics_rows = []
            
            # Add input pool metrics
            input_pool = results.get("input_pool")
            if input_pool:
                all_metrics_rows.extend(
                    self._flatten_metrics_to_rows(input_pool, pool_type="Input Pool")
                )
            
            # Add competitor metrics
            for competitor in results.get("competitors", []):
                all_metrics_rows.extend(
                    self._flatten_metrics_to_rows(competitor, pool_type="Competitor")
                )
            
            # Create DataFrame and write
            if all_metrics_rows:
                df_metrics = pd.DataFrame(all_metrics_rows)
                df_metrics.to_excel(writer, sheet_name="all_metrics", index=False)
            
            # Sheet 2: Summary
            df_summary = self._create_summary_sheet(results, anchor_data)
            if not df_summary.empty:
                df_summary.to_excel(writer, sheet_name="Summary", index=False)
            
            # Sheet 3: Anchor Token Data (if provided)
            if anchor_df is not None and not anchor_df.empty:
                anchor_df.to_excel(writer, sheet_name="Anchor Token", index=False)
            elif anchor_data and anchor_data.get("markets"):
                # Fallback for old format
                df_anchor = pd.DataFrame(anchor_data["markets"])
                df_anchor.to_excel(writer, sheet_name="Anchor Token", index=False)
            
            # Apply formatting
            self._format_excel_sheets(writer)
        
        print(f"✅ Exported to Excel: {filepath}")
        return str(filepath)
        
    def export_to_csv(
        self,
        results: dict[str, Any],
        anchor_data: dict[str, Any] | None = None,
        filename: str | None = None,
    ) -> str:
        """
        Export metrics pipeline results to CSV file.
        
        Args:
            results: Output from MetricsPipeline.analyze_pool_with_competitors()
            anchor_data: Optional anchor token data
            filename: Optional custom filename
            
        Returns:
            Path to created CSV file
        """
        # Generate filename
        if not filename:
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            pool_addr = results.get("input_pool", {}).get("pool_address", "unknown")[:8]
            filename = f"pool_metrics_{pool_addr}_{timestamp}.csv"
        
        filepath = self.export_dir / filename
        
        # Flatten all metrics
        all_rows = []
        
        # Input pool
        input_pool = results.get("input_pool")
        if input_pool:
            all_rows.extend(self._flatten_metrics_to_rows(input_pool, "Input Pool"))
        
        # Competitors
        for comp in results.get("competitors", []):
            all_rows.extend(self._flatten_metrics_to_rows(comp, "Competitor"))
        
        # Create and save DataFrame
        if all_rows:
            df = pd.DataFrame(all_rows)
            df.to_csv(filepath, index=False)
            print(f"✅ Exported to CSV: {filepath}")
        else:
            print(f"⚠️  No data to export to CSV")
        
        return str(filepath)
        
    def _flatten_metrics_to_rows(
        self,
        pool_data: dict[str, Any],
        pool_type: str = "Input Pool",
    ) -> list[dict[str, Any]]:
        """
        Convert nested metrics dict to flat list of rows.
        
        Args:
            pool_data: Pool data with metrics from pipeline
            pool_type: "Input Pool" or "Competitor"
            
        Returns:
            List of row dictionaries ready for DataFrame
            
        Example structure:
            [
                {
                    "pool_name": "50WETH-50USDC",
                    "pool_address": "0x123...",
                    "pool_type": "Input Pool",
                    "dex": "Balancer",
                    "blockchain": "ethereum",
                    "metric_group": "Demand / Usage Metrics",
                    "metric_name": "Total Traded Volume",
                    "metric_value": 1234567.89,
                    "metric_unit": "USD",
                    "timestamp": "2024-02-11 15:30:00 UTC"
                },
                ...
            ]
        """
        rows = []
        timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        
        pool_name = pool_data.get("pool_name", "Unknown")
        pool_address = pool_data.get("pool_address", "N/A")
        dex = pool_data.get("dex", "Unknown")
        blockchain = pool_data.get("blockchain", "ethereum")
        
        # Get metrics dict
        metrics = pool_data.get("metrics", {})
        
        # Iterate through metric groups
        for group_key, group_metrics in metrics.items():
            # Get display name for group
            group_name = METRIC_GROUP_NAMES.get(group_key, group_key)
            
            if not isinstance(group_metrics, dict):
                continue
            
            # Check if this metric group has an error
            if "error" in group_metrics:
                row = {
                    "pool_name": pool_name,
                    "pool_address": pool_address,
                    "pool_type": pool_type,
                    "dex": dex,
                    "blockchain": blockchain,
                    "metric_group": group_name,
                    "metric_name": "Error",
                    "metric_value": str(group_metrics["error"]),
                    "metric_unit": "Error",
                    "timestamp": timestamp,
                }
                rows.append(row)
                continue
            
            # Iterate through metrics in group
            for metric_name, metric_value in group_metrics.items():
                # Skip special keys like 'rows'
                if metric_name in ["rows", "error"]:
                    continue
                
                # Determine unit
                unit = self._infer_unit(metric_name, metric_value)
                
                # Create row
                row = {
                    "pool_name": pool_name,
                    "pool_address": pool_address,
                    "pool_type": pool_type,
                    "dex": dex,
                    "blockchain": blockchain,
                    "metric_group": group_name,
                    "metric_name": self._format_metric_name(metric_name),
                    "metric_value": metric_value,
                    "metric_unit": unit,
                    "timestamp": timestamp,
                }
                rows.append(row)
        
        return rows
    
    def _infer_unit(self, metric_name: str, value: Any) -> str:
        """
        Infer unit from metric name.
        
        Args:
            metric_name: Name of the metric
            value: Value of the metric
            
        Returns:
            Inferred unit string
        """
        name_lower = metric_name.lower()
        
        if "volume" in name_lower or "tvl" in name_lower or "liquidity" in name_lower:
            return "USD"
        elif "count" in name_lower or "trades" in name_lower:
            return "Count"
        elif "fee" in name_lower or "apr" in name_lower or "apy" in name_lower:
            return "%"
        elif "ratio" in name_lower or "efficiency" in name_lower:
            return "Ratio"
        elif "percent" in name_lower or "change" in name_lower:
            return "%"
        else:
            return "Value"
    
    def _format_metric_name(self, name: str) -> str:
        """
        Convert snake_case to Title Case.
        
        Args:
            name: Metric name in snake_case
            
        Returns:
            Formatted metric name
        """
        return name.replace("_", " ").title()
        
    def _create_summary_sheet(self, results: dict[str, Any], anchor_data: dict[str, Any] | None = None) -> pd.DataFrame:
        """
        Create summary sheet with high-level stats per pool and anchor token.
        
        Args:
            results: Pipeline results
            anchor_data: Optional anchor token summary data
            
        Returns:
            DataFrame with summary statistics
            
        Columns:
            - pool_name
            - pool_address  
            - pool_type
            - dex
            - blockchain
            - total_metrics
            - metric_groups_count
            - anchor_token (if provided)
            - anchor_best_apy (if provided)
        """
        summary_rows = []
        
        # Input pool
        input_pool = results.get("input_pool")
        if input_pool:
            metrics = input_pool.get("metrics", {})
            total_metrics = sum(
                len([k for k in v.keys() if k not in ["rows", "error"]])
                for v in metrics.values()
                if isinstance(v, dict)
            )
            
            row = {
                "pool_name": input_pool.get("pool_name", "Unknown"),
                "pool_address": input_pool.get("pool_address", "N/A"),
                "pool_type": "Input Pool",
                "dex": input_pool.get("dex", "Unknown"),
                "blockchain": input_pool.get("blockchain", "ethereum"),
                "total_metrics": total_metrics,
                "metric_groups_count": len(metrics),
            }
            
            # Add anchor token info if available
            if anchor_data:
                row["anchor_token"] = anchor_data.get("token_symbol", "N/A")
                row["anchor_token_address"] = anchor_data.get("token_address", "N/A")
                stats = anchor_data.get("stats", {})
                row["anchor_total_markets"] = stats.get("total_markets", 0)
                row["anchor_avg_apy"] = stats.get("avg_apy", 0)
                row["anchor_max_apy"] = stats.get("max_apy", 0)
                if "cumulative_volume" in stats:
                    row["anchor_30d_volume_usd"] = stats.get("cumulative_volume", 0)
            
            summary_rows.append(row)
        
        # Competitors
        for i, competitor in enumerate(results.get("competitors", []), 1):
            metrics = competitor.get("metrics", {})
            total_metrics = sum(
                len([k for k in v.keys() if k not in ["rows", "error"]])
                for v in metrics.values()
                if isinstance(v, dict)
            )
            
            summary_rows.append({
                "pool_name": competitor.get("pool_name", "Unknown"),
                "pool_address": competitor.get("pool_address", "N/A"),
                "pool_type": f"Competitor #{i}",
                "dex": competitor.get("dex", "Unknown"),
                "blockchain": competitor.get("blockchain", "ethereum"),
                "total_metrics": total_metrics,
                "metric_groups_count": len(metrics),
            })
        
        return pd.DataFrame(summary_rows)
        
    def _format_excel_sheets(self, writer: pd.ExcelWriter) -> None:
        """
        Apply formatting to Excel sheets (column widths, headers).
        
        Args:
            writer: pandas ExcelWriter object
        """
        # Get workbook for openpyxl
        try:
            from openpyxl.styles import Font, PatternFill
            
            workbook = writer.book
            
            # Format each sheet
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    
                    # Set column width with some padding
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Format header row
                if worksheet.max_row > 0:
                    for cell in worksheet[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        except Exception as e:
            # Formatting is optional, don't fail if it doesn't work
            print(f"⚠️  Could not apply Excel formatting: {e}")
    
    def export_simple_pool_metrics(
        self,
        pool_data: dict[str, Any],
        metrics_data: dict[str, Any],
        anchor_data: dict[str, Any] | None = None,
        anchor_df: Any | None = None,
        format: str = "excel",
        filename: str | None = None,
    ) -> dict[str, str]:
        """
        Export simplified pool metrics (from MetricsCalculator, not MetricsPipeline).
        
        This is an adapter for main.py which uses MetricsCalculator instead of
        MetricsPipeline. It converts the simple metrics structure to a format
        compatible with the export methods.
        
        Args:
            pool_data: Raw pool data from BalancerAPI
            metrics_data: Formatted metrics from MetricsCalculator.format_metrics_for_email()
            anchor_data: Optional anchor token summary data
            anchor_df: Optional pandas DataFrame with full anchor token data
            format: "excel", "csv", or "both"
            filename: Optional base filename (without extension)
            
        Returns:
            Dictionary with export file paths: {"excel": path, "csv": path, "anchor_csv": path}
            
        Example:
            >>> exporter = DataExporter()
            >>> pool_data = await api.get_current_pool_data(pool_address)
            >>> metrics = await calculator.calculate_pool_metrics(pool_address)
            >>> metrics_data = calculator.format_metrics_for_email(metrics, pool_data)
            >>> files = exporter.export_simple_pool_metrics(
            ...     pool_data, metrics_data, format="both"
            ... )
            >>> print(files)
            {"excel": "exports/PoolReport_0x3de27_USDC_20260211_143022.xlsx",
             "csv": "exports/PoolMetrics_0x3de27_20260211_143022.csv",
             "anchor_csv": "exports/anchor_token_USDC_0xa0b869_20260211_143022.csv"}
        """
        # Generate base filename if not provided
        if not filename:
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            pool_addr_short = metrics_data.get("pool_address", "unknown")[:10]
            
            # Include anchor token in base filename if available
            if anchor_data and anchor_data.get("token_symbol"):
                token_symbol = anchor_data["token_symbol"]
                filename = f"PoolReport_{pool_addr_short}_{token_symbol}_{timestamp}"
            else:
                filename = f"PoolReport_{pool_addr_short}_{timestamp}"
        
        # Convert metrics_data to MetricsPipeline-compatible format
        converted_results = self._convert_simple_metrics_to_pipeline_format(
            pool_data, metrics_data
        )
        
        # Export using existing methods
        export_files = {}
        
        if format in ["excel", "both"]:
            excel_filename = f"{filename}.xlsx"
            excel_path = self.export_to_excel(
                converted_results,
                anchor_data=anchor_data,
                anchor_df=anchor_df,
                filename=excel_filename
            )
            export_files["excel"] = excel_path
            
        if format in ["csv", "both"]:
            # Separate CSV files for pool metrics and anchor token
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            pool_addr_short = metrics_data.get("pool_address", "unknown")[:10]
            
            # Pool metrics CSV
            csv_filename = f"PoolMetrics_{pool_addr_short}_{timestamp}.csv"
            csv_path = self.export_to_csv(
                converted_results,
                filename=csv_filename
            )
            export_files["csv"] = csv_path
            
        return export_files
    
    def _convert_simple_metrics_to_pipeline_format(
        self,
        pool_data: dict[str, Any],
        metrics_data: dict[str, Any],
    ) -> dict[str, Any]:
        """
        Convert MetricsCalculator output to MetricsPipeline format.
        
        This is a lossy conversion - we can only export the simple metrics
        that MetricsCalculator provides, not the full 8 Dune metric groups.
        
        Args:
            pool_data: Raw pool data
            metrics_data: Formatted metrics from MetricsCalculator
            
        Returns:
            Dictionary in MetricsPipeline format (limited data)
        """
        # Extract basic pool info
        pool_name = metrics_data.get("pool_name", "Unknown Pool")
        pool_address = metrics_data.get("pool_address", "")
        
        # Create simplified metric groups using available data
        # Note: We don't have all 8 Dune groups, so create limited structure
        simplified_metrics = {
            "basic_metrics": {
                "tvl_current": metrics_data.get("tvl_current"),
                "tvl_15_days_ago": metrics_data.get("tvl_15_days_ago"),
                "tvl_change_percent": metrics_data.get("tvl_change_percent"),
                "volume_15_days": metrics_data.get("volume_15_days"),
                "volume_change_percent": metrics_data.get("volume_change_percent"),
                "fees_15_days": metrics_data.get("fees_15_days"),
                "fees_change_percent": metrics_data.get("fees_change_percent"),
                "apr_current": metrics_data.get("apr_current"),
                "swap_fee": metrics_data.get("swap_fee"),
            },
            "phase_3_hold_vs_pool": metrics_data.get("hold_vs_pool"),
            "phase_4_parameter_changes": metrics_data.get("parameter_changes"),
        }
        
        # Create MetricsPipeline-compatible structure
        return {
            "input_pool": {
                "pool_name": pool_name,
                "pool_address": pool_address,
                "pool_type": pool_data.get("poolType", "Unknown"),
                "blockchain": pool_data.get("_blockchain", "ethereum"),
                "metrics": simplified_metrics,
            },
            "competitors": [],  # No competitor data from MetricsCalculator
        }
    
    def export_multi_pool_metrics(
        self,
        multi_metrics: Any,  # MultiPoolMetrics type
        anchor_data: dict[str, Any] | None = None,
        anchor_df: Any | None = None,
        format: str = "excel",
    ) -> dict[str, str]:
        """
        Export multi-pool comparison metrics to Excel and/or CSV.
        
        Args:
            multi_metrics: MultiPoolMetrics object with pools and rankings
            anchor_data: Optional anchor token summary data
            anchor_df: Optional pandas DataFrame with full anchor token data
            format: "excel", "csv", or "both"
            
        Returns:
            Dictionary with export file paths: {"excel": path, "csv": path}
            
        Example:
            >>> exporter = DataExporter()
            >>> multi_metrics = await calculator.calculate_multi_pool_metrics(pool_addresses)
            >>> files = exporter.export_multi_pool_metrics(
            ...     multi_metrics, anchor_data, format="both"
            ... )
            >>> print(files)
            {"excel": "exports/PoolReport_3pools_USDC_20260211_143022.xlsx",
             "csv": "exports/PoolMetrics_3pools_20260211_143022.csv"}
        """
        # Generate base filename
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        pool_count = len(multi_metrics.pools)
        token_symbol = anchor_data.get("token_symbol", "TOKEN") if anchor_data else "TOKEN"
        
        export_files = {}
        
        if format in ["excel", "both"]:
            filename = f"PoolReport_{pool_count}pools_{token_symbol}_{timestamp}.xlsx"
            excel_path = self._create_multi_pool_excel(
                multi_metrics, anchor_data, anchor_df, filename
            )
            export_files["excel"] = excel_path
        
        if format in ["csv", "both"]:
            filename = f"PoolMetrics_{pool_count}pools_{timestamp}.csv"
            csv_path = self._create_multi_pool_csv(multi_metrics, filename)
            export_files["csv"] = csv_path
        
        return export_files
    
    def _create_multi_pool_excel(
        self,
        multi_metrics: Any,
        anchor_data: dict[str, Any] | None,
        anchor_df: Any | None,
        filename: str,
    ) -> str:
        """
        Create Excel workbook with multiple tabs for multi-pool data.
        
        Tabs:
        1. Pool Comparison - All pools with key metrics
        2. Rankings - Top 3 by volume, TVL, and custom metrics
        3. Summary - Totals and weighted averages with anchor token info
        4. Anchor Token - Full anchor token data (if provided)
        """
        filepath = self.export_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Tab 1: Pool Comparison
            df_pools = self._multi_pools_to_dataframe(multi_metrics.pools)
            df_pools.to_excel(writer, sheet_name="Pool Comparison", index=False)
            
            # Tab 2: Rankings
            df_rankings = self._rankings_to_dataframe(multi_metrics)
            df_rankings.to_excel(writer, sheet_name="Rankings", index=False)
            
            # Tab 3: Summary
            df_summary = self._multi_pool_summary(multi_metrics, anchor_data)
            df_summary.to_excel(writer, sheet_name="Summary", index=False)
            
            # Tab 4: Anchor Token (if provided)
            if anchor_df is not None and not anchor_df.empty:
                anchor_df.to_excel(writer, sheet_name="Anchor Token", index=False)
            
            # Apply formatting
            self._format_excel_sheets(writer)
        
        print(f"✅ Exported multi-pool Excel: {filepath}")
        return str(filepath)
    
    def _create_multi_pool_csv(
        self,
        multi_metrics: Any,
        filename: str,
    ) -> str:
        """Create flat CSV with all pool metrics."""
        filepath = self.export_dir / filename
        
        # Convert pools to DataFrame
        df = self._multi_pools_to_dataframe(multi_metrics.pools)
        df.to_csv(filepath, index=False)
        
        print(f"✅ Exported multi-pool CSV: {filepath}")
        return str(filepath)
    
    def _multi_pools_to_dataframe(self, pools: list[Any]) -> pd.DataFrame:
        """
        Convert list of PoolMetrics to flat DataFrame.
        
        Args:
            pools: List of PoolMetrics objects
            
        Returns:
            DataFrame with one row per pool
        """
        rows = []
        for pool in pools:
            row = {
                "pool_name": pool.pool_name,
                "pool_address": pool.pool_address,
                "pool_type": pool.pool_type,
                "boosted": pool.boosted_type,
                "pool_url": pool.pool_url,
                "tvl_current_usd": pool.tvl_current,
                "tvl_15d_ago_usd": pool.tvl_15_days_ago,
                "tvl_change_%": pool.tvl_change_percent,
                "volume_15d_usd": pool.volume_15_days,
                "volume_change_%": pool.volume_change_percent,
                "fees_15d_usd": pool.fees_15_days,
                "fees_change_%": pool.fees_change_percent,
                "apr_current_%": pool.apr_current * 100 if pool.apr_current else None,
                "swap_fee_%": pool.swap_fee * 100,
                "is_core_pool": pool.is_core_pool,
            }
            
            # Add pool-type specific metrics
            if pool.boosted_apr is not None:
                row["boosted_apr_%"] = pool.boosted_apr * 100
                row["boosted_apr_15d_ago_%"] = pool.boosted_apr_15d_ago * 100 if pool.boosted_apr_15d_ago else None
            
            if pool.surge_fees is not None:
                row["surge_fees"] = pool.surge_fees
                row["surge_fees_15d_ago"] = pool.surge_fees_15d_ago
            
            if pool.rebalance_count_15d is not None:
                row["rebalance_count_15d"] = pool.rebalance_count_15d
            
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _rankings_to_dataframe(self, multi_metrics: Any) -> pd.DataFrame:
        """
        Convert rankings data to DataFrame.
        
        Creates a flat structure with all rankings (volume, TVL, custom).
        """
        rows = []
        
        # Volume rankings
        for rank, (name, volume, pct, url) in enumerate(multi_metrics.top_3_by_volume, 1):
            rows.append({
                "ranking_type": "Volume (15d)",
                "rank": rank,
                "pool_name": name,
                "value": volume,
                "percentage": pct,
                "pool_url": url,
            })
        
        # TVL increase rankings
        for rank, (name, tvl_increase, pct, url) in enumerate(multi_metrics.top_3_by_tvl, 1):
            rows.append({
                "ranking_type": "TVL Increase",
                "rank": rank,
                "pool_name": name,
                "value": tvl_increase,
                "percentage": pct,
                "pool_url": url,
            })
        
        # Custom rankings
        for metric_name, ranking in multi_metrics.custom_rankings.items():
            display_name = metric_name.replace("_", " ").title()
            for rank, (name, value, url) in enumerate(ranking, 1):
                rows.append({
                    "ranking_type": display_name,
                    "rank": rank,
                    "pool_name": name,
                    "value": value,
                    "percentage": None,
                    "pool_url": url,
                })
        
        return pd.DataFrame(rows)
    
    def _multi_pool_summary(
        self, 
        multi_metrics: Any,
        anchor_data: dict[str, Any] | None,
    ) -> pd.DataFrame:
        """
        Create summary statistics for multi-pool comparison.
        
        Includes totals, averages, and anchor token info.
        """
        metrics = []
        values = []
        
        # Pool statistics
        metrics.append("Total Pools")
        values.append(len(multi_metrics.pools))
        
        # Count boosted pools
        boosted_pools = [p for p in multi_metrics.pools if p.boosted_type]
        metrics.append("Boosted Pools")
        values.append(len(boosted_pools))
        
        metrics.append("Total TVL (All Pools)")
        values.append(sum(p.tvl_current for p in multi_metrics.pools))
        
        metrics.append("Total Volume (15d)")
        values.append(sum(p.volume_15_days for p in multi_metrics.pools))
        
        metrics.append("Total Fees (15d)")
        values.append(multi_metrics.total_fees)
        
        metrics.append("Weighted APR (%)")
        values.append(multi_metrics.total_apr * 100)
        
        # Anchor token statistics (if available)
        if anchor_data:
            metrics.append("--- Anchor Token ---")
            values.append("")
            
            metrics.append("Token Symbol")
            values.append(anchor_data.get("token_symbol", "N/A"))
            
            metrics.append("Token Address")
            values.append(anchor_data.get("token_address", "N/A"))
            
            stats = anchor_data.get("stats", {})
            
            metrics.append("Lending Markets Found")
            values.append(stats.get("total_markets", 0))
            
            if "avg_apy" in stats:
                metrics.append("Average APY (%)")
                values.append(stats.get("avg_apy", 0))
            
            if "max_apy" in stats:
                metrics.append("Maximum APY (%)")
                values.append(stats.get("max_apy", 0))
            
            if "cumulative_volume" in stats:
                metrics.append("30-Day Volume (USD)")
                values.append(stats.get("cumulative_volume", 0))
        
        return pd.DataFrame({
            "metric": metrics,
            "value": values,
        })

    def cleanup_old_exports(self, max_age_hours: int = 24) -> int:
        """
        Delete export files older than specified hours.
        
        Args:
            max_age_hours: Delete files older than this many hours
            
        Returns:
            Number of files deleted
        """
        deleted = 0
        cutoff_time = datetime.now(timezone.utc) - timedelta(hours=max_age_hours)
        
        for filepath in self.export_dir.glob("*.*"):
            if filepath.is_file():
                file_mtime = datetime.fromtimestamp(filepath.stat().st_mtime, tz=timezone.utc)
                if file_mtime < cutoff_time:
                    try:
                        filepath.unlink()
                        deleted += 1
                    except Exception as e:
                        print(f"⚠️  Failed to delete {filepath.name}: {e}")
        
        if deleted > 0:
            print(f"🗑️  Cleaned up {deleted} old export files")
        
        return deleted
